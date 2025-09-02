"""
BOM 원가 계산기 - SharePoint 연동 버전
SharePoint에서 BOM 데이터를 자동으로 가져와서 계산하는 버전
"""

import streamlit as st
import pandas as pd
import numpy as np
import io
import requests
import json
from typing import Dict, Optional, Tuple
from datetime import datetime
import time
import warnings

# 경고 필터링
warnings.filterwarnings('ignore')

# 선택적 import (없어도 동작)
try:
    from tqdm import stqdm
    HAS_PROGRESS = True
except ImportError:
    HAS_PROGRESS = False

try:
    import plotly.express as px
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False

class Config:
    """설정 클래스"""
    MAX_FILE_SIZE_MB = 100
    REQUIRED_BOM_COLS = ['생산품목코드', '생산품목명', '소모품목코드', '소모품목명', '소요량']
    TEST_ITEM_CODE = '99701'

class SharePointClient:
    """SharePoint 연동 클래스"""
    
    def __init__(self):
        self.access_token = None
        self.token_expires_at = 0
        
    def get_access_token(self) -> Optional[str]:
        """Azure AD 토큰 획득"""
        try:
            # Streamlit secrets에서 설정 읽기
            tenant_id = st.secrets["sharepoint"]["tenant_id"]
            client_id = st.secrets["sharepoint"]["client_id"]
            client_secret = st.secrets["sharepoint"]["client_secret"]
            
            # 토큰이 유효한지 확인 (5분 여유)
            if self.access_token and time.time() < self.token_expires_at - 300:
                return self.access_token
            
            # 새 토큰 요청
            token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
            
            token_data = {
                'grant_type': 'client_credentials',
                'client_id': client_id,
                'client_secret': client_secret,
                'scope': 'https://graph.microsoft.com/.default'
            }
            
            response = requests.post(token_url, data=token_data)
            response.raise_for_status()
            
            token_info = response.json()
            self.access_token = token_info['access_token']
            self.token_expires_at = time.time() + token_info.get('expires_in', 3600)
            
            st.success("🔑 SharePoint 인증 성공!")
            return self.access_token
            
        except Exception as e:
            st.error(f"❌ SharePoint 인증 실패: {e}")
            return None
    
    def get_site_id(self, site_name: str) -> Optional[str]:
        """사이트 ID 획득"""
        try:
            token = self.get_access_token()
            if not token:
                return None
            
            headers = {
                'Authorization': f'Bearer {token}',
                'Accept': 'application/json'
            }
            
            # 사이트 정보 가져오기
            site_url = f"https://graph.microsoft.com/v1.0/sites/goremi.sharepoint.com:/sites/{site_name}"
            response = requests.get(site_url, headers=headers)
            response.raise_for_status()
            
            site_info = response.json()
            return site_info['id']
            
        except Exception as e:
            st.error(f"❌ 사이트 ID 획득 실패: {e}")
            return None
    
    def download_file_from_sharepoint(self, file_url: str) -> Optional[bytes]:
        """SharePoint 파일 다운로드 (개선된 버전)"""
        try:
            token = self.get_access_token()
            if not token:
                return None
            
            site_name = st.secrets["sharepoint_files"]["site_name"]
            file_name = st.secrets["sharepoint_files"]["file_name"]
            
            site_id = self.get_site_id(site_name)
            if not site_id:
                return None
            
            headers = {
                'Authorization': f'Bearer {token}',
                'Accept': 'application/json'
            }
            
            st.info(f"🔍 파일 검색 중: {file_name}")
            
            # 방법 1: 드라이브별 파일 검색
            drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
            drives_response = requests.get(drives_url, headers=headers)
            drives_response.raise_for_status()
            
            drives = drives_response.json()['value']
            if not drives:
                st.error("❌ SharePoint 드라이브를 찾을 수 없습니다")
                return None
            
            file_item = None
            drive_id = None
            
            # 모든 드라이브에서 파일 검색
            for drive in drives:
                try:
                    drive_id = drive['id']
                    st.info(f"🔍 드라이브 '{drive['name']}'에서 검색 중...")
                    
                    # 방법 1-1: 검색 API 사용
                    search_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/search(q='{file_name}')"
                    search_response = requests.get(search_url, headers=headers)
                    
                    if search_response.status_code == 200:
                        search_results = search_response.json().get('value', [])
                        if search_results:
                            file_item = search_results[0]
                            st.info(f"✅ 파일 발견: {file_item['name']}")
                            break
                    
                    # 방법 1-2: 루트 디렉터리 직접 조회
                    if not file_item:
                        root_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/children"
                        root_response = requests.get(root_url, headers=headers)
                        
                        if root_response.status_code == 200:
                            root_files = root_response.json().get('value', [])
                            for f in root_files:
                                if f['name'].lower() == file_name.lower():
                                    file_item = f
                                    st.info(f"✅ 루트에서 파일 발견: {file_item['name']}")
                                    break
                    
                    if file_item:
                        break
                        
                except Exception as drive_error:
                    st.warning(f"⚠️ 드라이브 '{drive['name']}' 검색 실패: {drive_error}")
                    continue
            
            if not file_item:
                st.error(f"❌ 파일 '{file_name}'을 찾을 수 없습니다")
                return None
            
            st.info(f"📁 파일 정보: {file_item['name']} (크기: {file_item.get('size', 0):,} bytes)")
            
            # 방법 2: 다운로드 URL 생성 (여러 방법 시도)
            download_url = None
            
            # 방법 2-1: @microsoft.graph.downloadUrl 사용
            if '@microsoft.graph.downloadUrl' in file_item:
                download_url = file_item['@microsoft.graph.downloadUrl']
                st.info("🔗 downloadUrl 방식 사용")
            
            # 방법 2-2: /content 엔드포인트 사용
            elif 'id' in file_item:
                download_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_item['id']}/content"
                st.info("🔗 content 엔드포인트 방식 사용")
            
            # 방법 2-3: 직접 경로로 접근
            else:
                download_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{file_name}:/content"
                st.info("🔗 직접 경로 방식 사용")
            
            if not download_url:
                st.error("❌ 다운로드 URL을 생성할 수 없습니다")
                return None
            
            # 파일 다운로드 실행
            st.info("📥 파일 다운로드 시작...")
            
            # content 엔드포인트는 Authorization 헤더 필요
            if 'graph.microsoft.com' in download_url and '/content' in download_url:
                download_headers = headers.copy()
            else:
                download_headers = {}  # downloadUrl은 이미 인증된 URL
            
            file_response = requests.get(download_url, headers=download_headers)
            file_response.raise_for_status()
            
            if len(file_response.content) == 0:
                st.error("❌ 다운로드된 파일이 비어있습니다")
                return None
            
            st.success(f"✅ 파일 다운로드 완료: {len(file_response.content):,} bytes")
            return file_response.content
            
        except requests.exceptions.HTTPError as e:
            st.error(f"❌ HTTP 오류: {e.response.status_code} - {e.response.text}")
            return None
        except Exception as e:
            st.error(f"❌ SharePoint 파일 다운로드 실패: {str(e)}")
            st.error(f"🔍 디버그 정보: {type(e).__name__}")
            return None
    
    def get_file_info(self, file_url: str) -> Optional[Dict]:
        """SharePoint 파일 정보 조회"""
        try:
            token = self.get_access_token()
            if not token:
                return None
            
            site_name = st.secrets["sharepoint_files"]["site_name"]
            file_name = st.secrets["sharepoint_files"]["file_name"]
            
            site_id = self.get_site_id(site_name)
            if not site_id:
                return None
            
            headers = {
                'Authorization': f'Bearer {token}',
                'Accept': 'application/json'
            }
            
            # 드라이브 정보 가져오기
            drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
            drives_response = requests.get(drives_url, headers=headers)
            drives_response.raise_for_status()
            
            drives = drives_response.json()['value']
            if drives:
                drive_id = drives[0]['id']
                
                # 파일 검색
                search_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/search(q='{file_name}')"
                search_response = requests.get(search_url, headers=headers)
                search_response.raise_for_status()
                
                search_results = search_response.json()['value']
                if search_results:
                    file_item = search_results[0]
                    return {
                        'name': file_item['name'],
                        'size': file_item['size'],
                        'last_modified': file_item['lastModifiedDateTime'],
                        'created': file_item['createdDateTime']
                    }
            
            return None
            
        except Exception as e:
            st.error(f"❌ 파일 정보 조회 실패: {e}")
            return None

def validate_file_size(file_content: bytes, max_mb: int = 100) -> bool:
    """파일 크기 검증"""
    try:
        size_mb = len(file_content) / (1024 * 1024)
        if size_mb > max_mb:
            st.error(f"❌ 파일이 너무 큽니다: {size_mb:.1f}MB > {max_mb}MB")
            return False
        return True
    except Exception:
        return False

def safe_load_data(file_content: bytes, file_name: str, skiprows: int = 0) -> Optional[pd.DataFrame]:
    """안전한 파일 로딩"""
    try:
        file_obj = io.BytesIO(file_content)
        
        # 파일 형식별 로딩
        if file_name.lower().endswith('.csv'):
            # CSV 인코딩 시도
            for encoding in ['utf-8-sig', 'utf-8', 'cp949', 'euc-kr']:
                try:
                    file_obj.seek(0)
                    df = pd.read_csv(file_obj, skiprows=skiprows, encoding=encoding, dtype=str)
                    break
                except:
                    continue
            else:
                st.error("❌ CSV 파일 인코딩을 감지할 수 없습니다")
                return None
                
        elif file_name.lower().endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file_obj, skiprows=skiprows, dtype=str)
        else:
            st.error("❌ 지원하지 않는 파일 형식")
            return None
        
        # 헤더 문제 해결 (구매 데이터용)
        if 'purchase' in file_name.lower() or any('Unnamed:' in str(col) for col in df.columns):
            df = fix_purchase_data_headers(df)
        
        # 데이터 정제
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].astype(str).str.strip()
        
        # 빈 행/열 제거
        df = df.dropna(how='all').dropna(axis=1, how='all')
        
        return df if not df.empty else None
        
    except Exception as e:
        st.error(f"❌ 파일 로딩 실패: {e}")
        return None

def fix_purchase_data_headers(df: pd.DataFrame) -> pd.DataFrame:
    """구매 데이터 헤더 문제 해결"""
    try:
        st.info("🔧 구매 데이터 헤더 문제를 감지했습니다. 자동으로 수정합니다...")
        
        # 첫 번째 행에서 실제 헤더 찾기
        potential_headers = None
        
        # 0행부터 3행까지 헤더 후보 검색
        for i in range(min(4, len(df))):
            row_values = df.iloc[i].fillna('').astype(str).tolist()
            row_text = ' '.join(row_values)
            
            # 헤더의 특징적인 키워드들 확인
            header_keywords = ['일자', '품목코드', '품목명', '단가', '수량', '거래처']
            keyword_count = sum(1 for keyword in header_keywords if keyword in row_text)
            
            if keyword_count >= 3:  # 3개 이상의 키워드가 있으면 헤더로 판단
                potential_headers = row_values
                header_row_idx = i
                st.info(f"📋 {i}행에서 실제 헤더를 발견했습니다: {keyword_count}개 키워드 매칭")
                break
        
        if potential_headers:
            # 헤더 정리 및 적용
            cleaned_headers = []
            for header in potential_headers:
                # 헤더명 정리
                header = str(header).strip()
                if header in ['', 'nan', 'None']:
                    # 빈 헤더는 이전 헤더 기반으로 생성
                    header = f"컬럼_{len(cleaned_headers)+1}"
                
                cleaned_headers.append(header)
            
            # 새로운 DataFrame 생성
            new_df = df.iloc[header_row_idx + 1:].copy()  # 헤더 다음 행부터 데이터
            new_df.columns = cleaned_headers[:len(new_df.columns)]  # 컬럼 수만큼만 헤더 적용
            new_df = new_df.reset_index(drop=True)
            
            # 빈 행 제거
            new_df = new_df.dropna(how='all')
            
            st.success(f"✅ 헤더 수정 완료: {len(cleaned_headers)}개 컬럼, {len(new_df)}행 데이터")
            st.write("**수정된 헤더:**", cleaned_headers[:10])  # 처음 10개만 표시
            
            return new_df
        
        else:
            st.warning("⚠️ 적절한 헤더를 찾을 수 없습니다. 원본 데이터를 사용합니다.")
            return df
        
    except Exception as e:
        st.error(f"❌ 헤더 수정 중 오류: {e}")
        return df

def validate_bom_data(df: pd.DataFrame) -> Tuple[bool, str]:
    """BOM 데이터 간단 검증"""
    try:
        # 필수 컬럼 확인
        missing_cols = [col for col in Config.REQUIRED_BOM_COLS if col not in df.columns]
        if missing_cols:
            return False, f"필수 컬럼 누락: {missing_cols}"
        
        # 데이터 존재 확인
        if df.empty:
            return False, "데이터가 비어있습니다"
        
        return True, "검증 통과"
        
    except Exception as e:
        return False, f"검증 오류: {e}"

def extract_purchase_prices(df: pd.DataFrame) -> Dict[str, float]:
    """구매 데이터에서 단가 추출 (헤더 문제 해결 포함)"""
    try:
        if df.empty:
            return {}
        
        st.write("📊 **구매 데이터 분석 중...**")
        st.write(f"- 원본 컬럼: {list(df.columns)[:5]}...")  # 처음 5개만 표시
        
        # 컬럼 자동 감지 (개선된 버전)
        date_col, item_col, price_col = None, None, None
        
        # 각 컬럼명을 분석하여 매칭
        for col in df.columns:
            col_str = str(col).lower()
            col_original = str(col)
            
            # 일자 컬럼 감지
            if not date_col:
                if any(keyword in col_str for keyword in ['일자', 'date', '날짜']) or '일자-no' in col_str:
                    date_col = col
                    st.info(f"📅 일자 컬럼 발견: '{col_original}'")
            
            # 품목코드 컬럼 감지  
            if not item_col:
                if '품목코드' in col_str:
                    item_col = col
                    st.info(f"🔖 품목코드 컬럼 발견: '{col_original}'")
            
            # 단가 컬럼 감지
            if not price_col:
                if '단가' in col_str and '공급' not in col_str and '총' not in col_str:
                    price_col = col
                    st.info(f"💰 단가 컬럼 발견: '{col_original}'")
        
        # 컬럼을 찾지 못한 경우 인덱스로 대체
        if not date_col and len(df.columns) > 0:
            date_col = df.columns[0]
            st.warning(f"⚠️ 일자 컬럼을 자동 설정: '{date_col}'")
            
        if not item_col and len(df.columns) > 1:
            item_col = df.columns[1]
            st.warning(f"⚠️ 품목코드 컬럼을 자동 설정: '{item_col}'")
            
        if not price_col:
            # 단가 관련 컬럼 우선 탐색
            for i, col in enumerate(df.columns):
                if i >= 3:  # 3번째 컬럼부터
                    sample_value = str(df[col].dropna().iloc[0] if not df[col].dropna().empty else '')
                    # 숫자로 변환 가능한 컬럼 찾기
                    try:
                        float(sample_value.replace(',', ''))
                        price_col = col
                        st.warning(f"⚠️ 단가 컬럼을 추정 설정: '{col}' (샘플값: {sample_value})")
                        break
                    except:
                        continue
            
            # 그래도 없으면 기본값
            if not price_col and len(df.columns) > 5:
                price_col = df.columns[5]
                st.warning(f"⚠️ 단가 컬럼을 기본 설정: '{price_col}'")
        
        if not all([date_col, item_col, price_col]):
            st.error(f"❌ 필수 컬럼을 찾을 수 없습니다. 일자: {date_col}, 품목코드: {item_col}, 단가: {price_col}")
            return {}
        
        st.success(f"✅ 컬럼 매핑 완료 - 일자: {date_col}, 품목코드: {item_col}, 단가: {price_col}")
        
        # 데이터 정제
        work_df = df[[date_col, item_col, price_col]].copy()
        work_df.columns = ['date', 'item_code', 'price']
        
        # 빈값 제거
        work_df = work_df.dropna()
        
        # 타입 변환
        work_df['item_code'] = work_df['item_code'].astype(str).str.strip()
        work_df['price'] = pd.to_numeric(work_df['price'], errors='coerce')
        
        # 유효한 데이터만
        work_df = work_df[
            (work_df['item_code'] != '') & 
            (work_df['item_code'] != 'nan') &
            (work_df['price'] > 0) & 
            (work_df['price'].notna())
        ]
        
        if work_df.empty:
            st.error("❌ 유효한 구매 데이터가 없습니다.")
            return {}
        
        # 날짜 처리 (간단하게)
        try:
            work_df['date_str'] = work_df['date'].astype(str).str.split('-').str[0]
            work_df['date_parsed'] = pd.to_datetime(work_df['date_str'], errors='coerce')
            work_df = work_df.dropna(subset=['date_parsed'])
            work_df = work_df.sort_values('date_parsed', ascending=False)
            st.info("📅 날짜순 정렬 완료")
        except:
            st.warning("⚠️ 날짜 정렬 실패, 원본 순서 유지")
        
        # 최신 단가 추출
        latest_prices = work_df.drop_duplicates(subset='item_code', keep='first')
        
        # 딕셔너리 변환
        price_dict = {}
        for _, row in latest_prices.iterrows():
            code = row['item_code']
            price = row['price']
            if pd.notna(price) and price > 0:
                price_dict[code] = float(price)
        
        st.write(f"**구매단가 샘플 (처음 5개):**")
        sample_items = list(price_dict.items())[:5]
        for code, price in sample_items:
            st.write(f"  • {code}: {price:,.0f}원")
        
        return price_dict
        
    except Exception as e:
        st.error(f"❌ 구매 단가 추출 오류: {e}")
        return {}

def clean_bom_data(df: pd.DataFrame) -> pd.DataFrame:
    """BOM 데이터 정제 (검증 없이)"""
    try:
        clean_df = df.copy()
        
        # 필수 컬럼 확인
        missing_cols = [col for col in Config.REQUIRED_BOM_COLS if col not in clean_df.columns]
        if missing_cols:
            st.error(f"❌ 필수 컬럼 누락: {missing_cols}")
            return pd.DataFrame()
        
        # 데이터 정제 - 검증 전에 타입 변환
        clean_df['생산품목코드'] = clean_df['생산품목코드'].astype(str).str.strip()
        clean_df['소모품목코드'] = clean_df['소모품목코드'].astype(str).str.strip()
        
        # 소요량 변환 (문제의 핵심)
        clean_df['소요량'] = pd.to_numeric(clean_df['소요량'], errors='coerce').fillna(0.0)
        
        # test 품목 제거
        before_count = len(clean_df)
        clean_df = clean_df[clean_df['소모품목코드'] != Config.TEST_ITEM_CODE]
        after_count = len(clean_df)
        
        if before_count != after_count:
            st.info(f"🧹 test 품목 제거: {before_count:,} → {after_count:,}행")
        
        # 유효하지 않은 데이터 제거
        clean_df = clean_df[
            (clean_df['생산품목코드'] != '') &
            (clean_df['소모품목코드'] != '') &
            (clean_df['생산품목코드'] != 'nan') &
            (clean_df['소모품목코드'] != 'nan') &
            (clean_df['소요량'] >= 0)
        ]
        
        return clean_df
        
    except Exception as e:
        st.error(f"❌ BOM 데이터 정제 오류: {e}")
        return pd.DataFrame()

def calculate_product_cost_with_reason(product_code: str, bom_df: pd.DataFrame, all_costs: Dict[str, float], cache: Dict[str, float]) -> Tuple[float, str]:
    """단일 제품 원가 계산 + 실패 이유 분석"""
    try:
        # 캐시 확인
        if product_code in cache:
            return cache[product_code], ""
        
        # BOM 구성요소 가져오기
        components = bom_df[bom_df['생산품목코드'] == product_code]
        
        if components.empty:
            cache[product_code] = 0.0
            return 0.0, "BOM 구성요소 없음"
        
        total_cost = 0.0
        missing_components = []
        zero_price_components = []
        invalid_quantity_components = []
        
        for _, comp in components.iterrows():
            comp_code = comp['소모품목코드']
            comp_name = comp['소모품목명']
            quantity = float(comp['소요량'])
            
            # 수량 검증
            if quantity <= 0:
                invalid_quantity_components.append(f"{comp_name}({comp_code})")
                continue
            
            # 부품 단가 찾기
            if comp_code in all_costs:
                unit_price = all_costs[comp_code]
                if unit_price <= 0:
                    zero_price_components.append(f"{comp_name}({comp_code})")
                    continue
            elif comp_code in bom_df['생산품목코드'].values:
                # 재귀 계산
                unit_price, _ = calculate_product_cost_with_reason(comp_code, bom_df, all_costs, cache)
                if unit_price <= 0:
                    missing_components.append(f"{comp_name}({comp_code})")
                    continue
                all_costs[comp_code] = unit_price
            else:
                missing_components.append(f"{comp_name}({comp_code})")
                continue
            
            total_cost += quantity * unit_price
        
        cache[product_code] = total_cost
        
        # 실패 이유 분석
        if total_cost == 0:
            reasons = []
            if missing_components:
                reasons.append(f"단가정보 없음: {', '.join(missing_components[:3])}{'...' if len(missing_components) > 3 else ''}")
            if zero_price_components:
                reasons.append(f"단가 0원: {', '.join(zero_price_components[:3])}{'...' if len(zero_price_components) > 3 else ''}")
            if invalid_quantity_components:
                reasons.append(f"수량 오류: {', '.join(invalid_quantity_components[:3])}{'...' if len(invalid_quantity_components) > 3 else ''}")
            
            failure_reason = " | ".join(reasons) if reasons else "알 수 없는 이유"
            return 0.0, failure_reason
        
        return total_cost, ""
        
    except Exception as e:
        return 0.0, f"계산 오류: {str(e)}"

def calculate_product_cost(product_code: str, bom_df: pd.DataFrame, all_costs: Dict[str, float], cache: Dict[str, float]) -> float:
    """단일 제품 원가 계산 (호환성 유지용)"""
    cost, _ = calculate_product_cost_with_reason(product_code, bom_df, all_costs, cache)
    return cost

def calculate_all_bom_costs(bom_df: pd.DataFrame, purchase_prices: Dict[str, float]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """전체 BOM 원가 계산"""
    try:
        start_time = time.time()
        
        # 데이터 정제
        clean_bom = clean_bom_data(bom_df)
        
        if clean_bom.empty:
            return pd.DataFrame(), pd.DataFrame()
        
        # 모든 생산품목
        all_products = clean_bom[['생산품목코드', '생산품목명']].drop_duplicates().reset_index(drop=True)
        
        st.info(f"📊 계산 대상: 생산품목 {len(all_products):,}개, 구매단가 {len(purchase_prices):,}개")
        
        # 전체 원가 딕셔너리
        all_costs = purchase_prices.copy()
        calc_cache = {}
        failure_reasons = {}  # 실패 이유 저장
        
        # 계산 실행
        results = []
        
        if HAS_PROGRESS:
            iterator = stqdm(all_products.iterrows(), total=len(all_products), desc="BOM 원가 계산")
        else:
            iterator = all_products.iterrows()
            progress_bar = st.progress(0)
        
        for idx, (_, product) in enumerate(iterator):
            product_code = product['생산품목코드']
            product_name = product['생산품목명']
            
            # 원가 계산
            calculated_cost, failure_reason = calculate_product_cost_with_reason(
                product_code, clean_bom, all_costs, calc_cache
            )
            
            if calculated_cost == 0 and failure_reason:
                failure_reasons[product_code] = failure_reason
            
            results.append({
                '생산품목코드': product_code,
                '생산품목명': product_name,
                '계산된단위원가': calculated_cost,
                '계산상태': '계산완료' if calculated_cost > 0 else '계산불가',
                '실패이유': failure_reasons.get(product_code, '')
            })
            
            # 진행률 업데이트
            if not HAS_PROGRESS:
                progress_bar.progress((idx + 1) / len(all_products))
        
        if not HAS_PROGRESS:
            progress_bar.empty()
        
        # 결과 DataFrame
        result_df = pd.DataFrame(results)
        
        # 상세 내역
        details_df = clean_bom.copy()
        details_df['부품단가'] = details_df['소모품목코드'].apply(lambda x: all_costs.get(x, 0.0))
        details_df['부품별원가'] = details_df['소요량'] * details_df['부품단가']
        
        elapsed = time.time() - start_time
        st.success(f"✅ 계산 완료! (소요시간: {elapsed:.1f}초)")
        
        return result_df, details_df
        
    except Exception as e:
        st.error(f"❌ BOM 원가 계산 실패: {e}")
        return pd.DataFrame(), pd.DataFrame()

def create_simple_chart(df: pd.DataFrame) -> None:
    """간단한 차트 생성"""
    if not HAS_PLOTLY or df.empty:
        return
    
    try:
        # 계산된 완제품만
        chart_data = df[df['계산된단위원가'] > 0]
        
        if len(chart_data) < 5:
            return
        
        # 상위 20개 제품 바차트
        top_items = chart_data.nlargest(20, '계산된단위원가')
        
        fig = px.bar(
            top_items,
            x='계산된단위원가',
            y='생산품목코드',
            orientation='h',
            title='원가 상위 20개 완제품',
            labels={'계산된단위원가': '원가 (원)', '생산품목코드': '제품코드'}
        )
        fig.update_layout(height=500, yaxis={'categoryorder': 'total ascending'})
        st.plotly_chart(fig, use_container_width=True)
        
    except Exception:
        pass  # 차트 실패해도 진행

def auto_adjust_column_width(worksheet, df: pd.DataFrame, start_row: int = 1):
    """엑셀 컬럼 너비 자동 조정"""
    try:
        from openpyxl.utils import get_column_letter
        
        for idx, column in enumerate(df.columns):
            column_letter = get_column_letter(idx + 1)
            
            # 헤더 길이
            header_length = len(str(column))
            
            # 데이터 최대 길이
            if not df.empty:
                max_data_length = df[column].astype(str).str.len().max()
            else:
                max_data_length = 0
            
            # 최적 너비 계산 (한글 고려)
            max_length = max(header_length, max_data_length)
            
            # 한글 문자가 많은 경우 추가 공간
            if any('\uac00' <= char <= '\ud7af' for char in str(column)):
                max_length = max_length * 1.3  # 한글은 30% 더
            
            # 최소/최대 너비 제한
            adjusted_width = min(max(max_length + 2, 10), 60)
            
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        st.warning(f"컬럼 너비 조정 실패: {e}")

def apply_excel_styling(worksheet, df: pd.DataFrame, sheet_title: str = ""):
    """엑셀 스타일링 적용"""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # 색상 정의
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        
        header_font = Font(color="FFFFFF", bold=True, size=11)
        title_font = Font(bold=True, size=14, color="1F4E79")
        data_font = Font(size=10)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 제목 스타일 (A1 셀)
        if sheet_title:
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = sheet_title
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 제목 셀 병합
            worksheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(df.columns))
        
        # 헤더 스타일 (2행 또는 1행)
        header_row = 2 if sheet_title else 1
        
        for col in range(1, len(df.columns) + 1):
            header_cell = worksheet.cell(row=header_row, column=col)
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = thin_border
        
        # 데이터 영역 스타일
        start_data_row = header_row + 1
        end_data_row = start_data_row + len(df) - 1
        
        for row in range(start_data_row, end_data_row + 1):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                
                # 숫자 컬럼 우측 정렬
                column_name = df.columns[col-1]
                if '원가' in column_name or '단가' in column_name or '수량' in column_name:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # 숫자 포맷
                    if isinstance(cell.value, (int, float)) and cell.value != 0:
                        cell.number_format = '#,##0'
        
        # 행 높이 조정
        worksheet.row_dimensions[header_row].height = 25
        for row in range(start_data_row, end_data_row + 1):
            worksheet.row_dimensions[row].height = 20
            
    except Exception as e:
        st.warning(f"스타일 적용 실패: {e}")

def export_to_excel(finished_goods: pd.DataFrame, all_results: pd.DataFrame, details: pd.DataFrame) -> bytes:
    """향상된 엑셀 내보내기 (자동 컬럼 조정 + 스타일링)"""
    try:
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. 완제품 원가 시트
            finished_display = finished_goods.copy()
            
            # 실패 이유가 있는 경우 컬럼명 변경
            if '실패이유' in finished_display.columns:
                finished_display = finished_display[['생산품목코드', '생산품목명', '계산된단위원가', '계산상태', '실패이유']]
                finished_display.columns = ['품목코드', '품목명', '단위원가(원)', '계산상태', '실패이유']
            else:
                finished_display = finished_display[['생산품목코드', '생산품목명', '계산된단위원가', '계산상태']]
                finished_display.columns = ['품목코드', '품목명', '단위원가(원)', '계산상태']
            
            # 데이터 쓰기 (제목용 공간 확보)
            finished_display.to_excel(writer, sheet_name='완제품원가', index=False, startrow=1)
            
            # 2. 전체 제품 원가 시트
            all_display = all_results.copy()
            if '실패이유' in all_display.columns:
                all_display = all_display[['생산품목코드', '생산품목명', '계산된단위원가', '계산상태', '실패이유']]
                all_display.columns = ['품목코드', '품목명', '단위원가(원)', '계산상태', '실패이유']
            
            all_display.to_excel(writer, sheet_name='전체제품원가', index=False, startrow=1)
            
            # 3. 상세 내역 시트
            details_display = details.copy()
            details_cols = ['생산품목코드', '생산품목명', '소모품목코드', '소모품목명', '소요량', '부품단가', '부품별원가']
            details_display = details_display[details_cols]
            details_display.columns = ['생산품목코드', '생산품목명', '부품코드', '부품명', '소요량', '부품단가(원)', '부품원가(원)']
            
            details_display.to_excel(writer, sheet_name='상세내역', index=False, startrow=1)
            
            # 각 시트에 스타일링 적용
            worksheets = [
                (writer.sheets['완제품원가'], finished_display, '완제품 BOM 원가 계산 결과'),
                (writer.sheets['전체제품원가'], all_display, '전체 제품 원가 계산 결과'),
                (writer.sheets['상세내역'], details_display, 'BOM 구성요소별 상세 원가 내역')
            ]
            
            for worksheet, df_data, title in worksheets:
                # 컬럼 너비 자동 조정
                auto_adjust_column_width(worksheet, df_data, start_row=2)
                
                # 스타일링 적용
                apply_excel_styling(worksheet, df_data, title)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"❌ 엑셀 생성 오류: {e}")
        return b''

def main():
    """메인 애플리케이션"""
    
    st.set_page_config(
        page_title="BOM 원가 계산기",
        page_icon="🏭",
        layout="wide"
    )
    
    st.title("🏭 BOM 원가 계산기 (SharePoint 연동 버전)")
    st.markdown("**🔗 SharePoint에서 BOM 데이터를 자동으로 가져와 계산하는 버전**")
    
    # 기능 상태 표시
    with st.sidebar:
        st.header("📊 시스템 상태")
        st.info(f"""
        **활성화된 기능:**
        - SharePoint 연동: ✅
        - 진행률 표시: {'✅' if HAS_PROGRESS else '❌'}
        - 시각화: {'✅' if HAS_PLOTLY else '❌'}
        """)
        
        # SharePoint 설정 확인
        try:
            tenant_id = st.secrets["sharepoint"]["tenant_id"]
            st.success("🔑 SharePoint 설정 확인됨")
        except:
            st.error("❌ SharePoint 설정이 없습니다")
    
    # SharePoint 클라이언트 초기화
    sharepoint_client = SharePointClient()
    
    # 파일 업로드 및 SharePoint 데이터 로딩
    st.header("1. 📁 데이터 소스")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📋 BOM 데이터 (SharePoint)")
        
        # SharePoint 파일 정보 표시
        if st.button("🔍 SharePoint BOM 파일 확인", type="secondary"):
            with st.spinner("SharePoint 연결 중..."):
                file_url = st.secrets["sharepoint_files"]["bom_file_url"]
                file_info = sharepoint_client.get_file_info(file_url)
                
                if file_info:
                    st.success("✅ SharePoint 파일 확인 완료!")
                    st.info(f"""
                    **파일 정보:**
                    - 파일명: {file_info['name']}
                    - 크기: {file_info['size']:,} bytes
                    - 수정일: {file_info['last_modified'][:19]}
                    """)
                else:
                    st.error("❌ SharePoint 파일을 찾을 수 없습니다")
        
        # 디버그 모드 추가
        debug_mode = st.checkbox("🔧 디버그 모드 활성화")
        
        if debug_mode:
            st.subheader("🔍 SharePoint 연결 디버깅")
            
            if st.button("🗂️ 사이트 드라이브 목록 보기"):
                with st.spinner("드라이브 정보 조회 중..."):
                    token = sharepoint_client.get_access_token()
                    if token:
                        site_name = st.secrets["sharepoint_files"]["site_name"]
                        site_id = sharepoint_client.get_site_id(site_name)
                        
                        if site_id:
                            headers = {'Authorization': f'Bearer {token}', 'Accept': 'application/json'}
                            drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
                            
                            try:
                                drives_response = requests.get(drives_url, headers=headers)
                                drives_response.raise_for_status()
                                drives = drives_response.json()['value']
                                
                                st.write("**발견된 드라이브:**")
                                for i, drive in enumerate(drives):
                                    st.write(f"{i+1}. {drive.get('name', 'Unnamed')} (ID: {drive['id']})")
                                    st.write(f"   - Type: {drive.get('driveType', 'Unknown')}")
                                    st.write(f"   - Owner: {drive.get('owner', {}).get('user', {}).get('displayName', 'Unknown')}")
                                
                            except Exception as e:
                                st.error(f"드라이브 조회 실패: {e}")
            
            if st.button("📁 파일 검색 테스트"):
                with st.spinner("파일 검색 테스트 중..."):
                    token = sharepoint_client.get_access_token()
                    if token:
                        site_name = st.secrets["sharepoint_files"]["site_name"]
                        file_name = st.secrets["sharepoint_files"]["file_name"]
                        site_id = sharepoint_client.get_site_id(site_name)
                        
                        if site_id:
                            headers = {'Authorization': f'Bearer {token}', 'Accept': 'application/json'}
                            drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
                            drives_response = requests.get(drives_url, headers=headers)
                            drives = drives_response.json()['value']
                            
                            for drive in drives:
                                drive_id = drive['id']
                                st.write(f"**드라이브 '{drive['name']}'에서 검색:**")
                                
                                # 검색 방법 1: Search API
                                try:
                                    search_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/search(q='{file_name}')"
                                    search_response = requests.get(search_url, headers=headers)
                                    if search_response.status_code == 200:
                                        results = search_response.json().get('value', [])
                                        st.write(f"  - Search API: {len(results)}개 결과")
                                        for r in results[:3]:  # 처음 3개만
                                            st.write(f"    * {r['name']} ({r.get('size', 0)} bytes)")
                                except Exception as e:
                                    st.write(f"  - Search API 실패: {e}")
                                
                                # 검색 방법 2: Root 디렉터리
                                try:
                                    root_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/children"
                                    root_response = requests.get(root_url, headers=headers)
                                    if root_response.status_code == 200:
                                        files = root_response.json().get('value', [])
                                        matching_files = [f for f in files if file_name.lower() in f['name'].lower()]
                                        st.write(f"  - Root 디렉터리: 전체 {len(files)}개 파일, 일치 {len(matching_files)}개")
                                        for f in matching_files:
                                            st.write(f"    * {f['name']} ({f.get('size', 0)} bytes)")
                                except Exception as e:
                                    st.write(f"  - Root 디렉터리 실패: {e}")
        
        # BOM 데이터 로딩 버튼
        if st.button("📥 SharePoint에서 BOM 데이터 가져오기", type="primary"):
            with st.spinner("SharePoint에서 BOM 데이터 다운로드 중..."):
                file_url = st.secrets["sharepoint_files"]["bom_file_url"]
                bom_content = sharepoint_client.download_file_from_sharepoint(file_url)
                
                if bom_content and validate_file_size(bom_content):
                    # 세션 상태에 저장
                    st.session_state['bom_content'] = bom_content
                    st.session_state['bom_filename'] = st.secrets["sharepoint_files"]["file_name"]
                    st.success("✅ BOM 데이터 로딩 완료!")
                else:
                    st.error("❌ BOM 데이터 로딩 실패")
        
        # 로딩된 BOM 데이터 표시
        if 'bom_content' in st.session_state:
            st.success(f"📋 BOM 데이터 준비됨: {st.session_state['bom_filename']}")
    
    with col2:
        st.subheader("💰 구매 데이터 (파일 업로드)")
        purchase_file = st.file_uploader("💰 구매 데이터 파일", type=['csv', 'xlsx', 'xls'], key="purchase")
    
    # 데이터 처리 및 계산
    if 'bom_content' in st.session_state and purchase_file:
        
        # 파일 크기 확인
        if not validate_file_size(purchase_file):
            st.stop()
        
        # 파일 로딩
        with st.spinner("📖 데이터 로딩 중..."):
            # BOM 데이터 (SharePoint)
            bom_df = safe_load_data(
                st.session_state['bom_content'], 
                st.session_state['bom_filename'], 
                skiprows=1
            )
            
            # 구매 데이터 (파일 업로드)
            purchase_df = safe_load_data(purchase_file.getvalue(), purchase_file.name)
        
        if bom_df is None or purchase_df is None:
            st.stop()
        
        # 간단한 검증
        bom_valid, bom_msg = validate_bom_data(bom_df)
        if not bom_valid:
            st.error(f"❌ BOM 데이터 오류: {bom_msg}")
            st.stop()
        
        # 데이터 미리보기
        st.header("2. 📋 데이터 미리보기")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📋 BOM 데이터 (SharePoint)")
            st.info(f"📊 {len(bom_df):,}행 × {len(bom_df.columns)}열")
            st.dataframe(bom_df.head(3), use_container_width=True)
            
        with col2:
            st.subheader("💰 구매 데이터")
            st.info(f"📊 {len(purchase_df):,}행 × {len(purchase_df.columns)}열")
            st.dataframe(purchase_df.head(3), use_container_width=True)
        
        # 원가 계산
        st.header("3. 🚀 BOM 원가 계산")
        
        if st.button("💪 원가 계산 시작!", type="primary", use_container_width=True):
            
            # 구매 단가 추출
            with st.spinner("💰 구매 단가 추출 중..."):
                purchase_prices = extract_purchase_prices(purchase_df)
            
            if not purchase_prices:
                st.error("❌ 구매 단가를 추출할 수 없습니다.")
                st.stop()
            
            st.success(f"✅ 구매 단가 추출 완료: {len(purchase_prices):,}개 품목")
            
            # BOM 원가 계산
            result_df, details_df = calculate_all_bom_costs(bom_df, purchase_prices)
            
            if result_df.empty:
                st.error("❌ BOM 원가 계산 실패")
                st.stop()
            
            # 완제품 필터링
            finished_goods = result_df[
                result_df['생산품목명'].str.contains('[완제품]', regex=False, na=False)
            ].copy()
            
            # 결과 표시
            st.header("4. 🎯 완제품 원가 결과")
            
            # 통계
            total = len(finished_goods)
            calculated = len(finished_goods[finished_goods['계산상태'] == '계산완료'])
            success_rate = (calculated / total * 100) if total > 0 else 0
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("🎯 전체 완제품", f"{total:,}개")
            with col2:
                st.metric("✅ 계산 성공", f"{calculated:,}개")
            with col3:
                st.metric("📊 성공률", f"{success_rate:.1f}%")
            
            # 결과 테이블
            if not finished_goods.empty:
                # 실패 이유 포함한 컬럼 구성
                if '실패이유' in finished_goods.columns:
                    display_df = finished_goods[['생산품목코드', '생산품목명', '계산된단위원가', '계산상태', '실패이유']].copy()
                    display_df.columns = ['품목코드', '품목명', '단위원가(원)', '상태', '실패이유']
                else:
                    display_df = finished_goods[['생산품목코드', '생산품목명', '계산된단위원가', '계산상태']].copy()
                    display_df.columns = ['품목코드', '품목명', '단위원가(원)', '상태']
                
                # 스타일링
                def highlight_rows(row):
                    if row['상태'] == '계산완료':
                        return ['background-color: #d4edda'] * len(row)
                    else:
                        return ['background-color: #f8d7da'] * len(row)
                
                styled_df = display_df.style.apply(highlight_rows, axis=1).format({
                    '단위원가(원)': '{:,.0f}'
                })
                
                st.dataframe(styled_df, use_container_width=True, height=400)
                
                # 실패 이유 분석 요약
                failed_items = finished_goods[finished_goods['계산상태'] == '계산불가']
                if not failed_items.empty and '실패이유' in failed_items.columns:
                    st.subheader("⚠️ 계산 실패 원인 분석")
                    
                    # 실패 이유별 통계
                    failure_stats = {}
                    for _, item in failed_items.iterrows():
                        reasons = item['실패이유'].split(' | ')
                        for reason in reasons:
                            main_reason = reason.split(':')[0] if ':' in reason else reason
                            failure_stats[main_reason] = failure_stats.get(main_reason, 0) + 1
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write("**실패 원인별 제품 수:**")
                        for reason, count in failure_stats.items():
                            st.write(f"• {reason}: {count}개")
                    
                    with col2:
                        st.write("**주요 개선 방향:**")
                        if '단가정보 없음' in failure_stats:
                            st.write("• 구매 데이터에 누락된 품목의 단가 정보 보완 필요")
                        if '단가 0원' in failure_stats:
                            st.write("• 단가가 0원인 품목의 정확한 단가 입력 필요")
                        if '수량 오류' in failure_stats:
                            st.write("• BOM 데이터의 소요량 정보 검토 및 수정 필요")
                
                # 원가 분석
                calculated_items = finished_goods[finished_goods['계산상태'] == '계산완료']
                if not calculated_items.empty and len(calculated_items) > 0:
                    avg_cost = calculated_items['계산된단위원가'].mean()
                    max_cost = calculated_items['계산된단위원가'].max()
                    min_cost = calculated_items[calculated_items['계산된단위원가'] > 0]['계산된단위원가'].min()
                    
                    st.subheader("📈 원가 통계")
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("💰 평균 원가", f"{avg_cost:,.0f}원")
                    with col2:
                        st.metric("📈 최고 원가", f"{max_cost:,.0f}원")
                    with col3:
                        st.metric("📉 최저 원가", f"{min_cost:,.0f}원")
                
                # 간단한 차트
                if HAS_PLOTLY:
                    st.subheader("📊 원가 분석 차트")
                    create_simple_chart(calculated_items)
            
            else:
                st.warning("⚠️ 완제품 데이터가 없습니다.")
            
            # 계산 실패 항목
            failed_items = finished_goods[finished_goods['계산상태'] == '계산불가']
            if not failed_items.empty:
                with st.expander(f"⚠️ 계산 실패 {len(failed_items):,}개 항목"):
                    st.dataframe(failed_items[['생산품목코드', '생산품목명']], use_container_width=True)
            
            # 결과 다운로드
            st.header("5. 📥 결과 다운로드")
            
            excel_data = export_to_excel(finished_goods, result_df, details_df)
            
            if excel_data:
                st.download_button(
                    label="📊 BOM 원가 계산 결과 다운로드 (Excel)",
                    data=excel_data,
                    file_name=f'BOM원가계산_SharePoint_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True
                )
            else:
                # CSV 대안
                csv_data = finished_goods.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📄 완제품 원가 결과 다운로드 (CSV)",
                    data=csv_data,
                    file_name=f'BOM원가계산_SharePoint_{datetime.now().strftime("%Y%m%d_%H%M")}.csv',
                    mime='text/csv',
                    use_container_width=True
                )
            
            st.balloons()
            st.success("🎉 BOM 원가 계산 완료!")
    
    else:
        st.info("👆 SharePoint에서 BOM 데이터를 가져오고, 구매 데이터 파일을 업로드해주세요.")
        
        # 간단한 사용법
        with st.expander("📖 사용법", expanded=True):
            st.markdown("""
            ### 🔗 SharePoint 연동 방식
            
            **1. BOM 데이터 (SharePoint 자동 연동):**
            - Azure AD 앱 등록 및 권한 설정 완료
            - SharePoint 파일을 자동으로 가져옴
            - 필수 컬럼: 생산품목코드, 생산품목명, 소모품목코드, 소모품목명, 소요량
            
            **2. 구매 데이터 (파일 업로드):**
            - CSV, Excel 파일 지원
            - 자동 헤더 감지 (일자, 품목코드, 단가)
            
            ### ⚡ 주요 특징
            - 🔗 **SharePoint 연동**: BOM 데이터 실시간 동기화
            - 🛡️ **Azure AD 인증**: 안전한 데이터 접근
            - 🎯 **자동화**: 수동 업로드 없이 최신 데이터 활용
            - 📊 **실시간 분석**: 진행률 및 결과 시각화
            
            ### 🔐 필수 설정 (Streamlit Secrets)
            - SharePoint 테넌트, 클라이언트 ID/Secret
            - Azure AD API 권한: Files.Read.All, Sites.Read.All
            """)

if __name__ == "__main__":
    main()
